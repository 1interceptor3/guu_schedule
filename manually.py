from bs4 import BeautifulSoup
import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.utils import coordinate_to_tuple
import requests
import os
from db import DataBaseSQLITE
import pprint


def make_obj(url):
    """
    Функция создаёт объект BeautifulSoup на основе URL

    :param url: URL страницы для создание объекта
    :type url: str
    :return: BeautifulSoup object
    :rtype: class BeautifulSoup
    """
    html = requests.request('GET', url)
    bs_obj = BeautifulSoup(html.text, 'html.parser')
    return bs_obj


def get_excel_file():
    """
    Функция ищет в текущей директории файлы с расширением .xlsx.
    В приоритете вернет которые не начинаются с символа '~', так как это временные файлы, которые создаются при
    открытии.

    :return: строка с названием excel файла в текущей директории
    """
    memory = None
    for i in os.listdir():
        if '.xlsx' in i and not i.startswith('~'):
            return i
        elif '.xlsx' in i:
            memory = i
    return memory


class Guu:
    link_schedule = 'https://guu.ru/%d1%81%d1%82%d1%83%d0%b4%d0%b5%d0%bd%d1%82%d0%b0%d0%bc/%d1%80%d0%b0%d1%81%d0%bf' \
                    '%d0%b8%d1%81%d0%b0%d0%bd%d0%b8%d0%b5-%d1%81%d0%b5%d1%81%d1%81%d0%b8%d0%b9/schedule'
    base_url = 'https://guu.ru'

    def __init__(self):
        self.db_obj = DataBaseSQLITE()

        # Проверка на необходимость обновления файла и на наличие файла
        if self.db_obj.need_to_update() or not get_excel_file():
            # Получаем новое название скачанного файла
            self.excel_file_name = self.download_file()
            print('New file', self.excel_file_name)
        else:
            self.excel_file_name = get_excel_file()
            print('Old file', self.excel_file_name)
            self.update_from_excel(self.excel_file_name)

    def download_file(self):
        # Удаляем старый файл
        for i in os.listdir():
            if '.xlsx' in i:
                os.remove(i)
        obj = make_obj(self.link_schedule)
        print('start downloading xlsx file')
        for i in obj.findAll('span', attrs={'class': 'doc-unit-title'}):
            if 'ОЗФО' in i.get_text():
                relative_path = i.parent.attrs['href']
                file_name = relative_path.split('/')[-1]
                html_file = requests.request('GET', self.base_url + relative_path)
                with open(file_name, 'wb') as file:
                    file.write(html_file.content)
                    print('xlsx file is downloaded')
                return file_name

    def update_from_excel(self, file_name):
        self.db_obj.delete_all_data()
        global inst_col, inst_row, prog_row  # Глобальные переменные для использования в разных циклах
        days_of_week = ('ПОНЕДЕЛЬНИК', 'ВТОРНИК', 'СРЕДА', 'ЧЕТВЕРГ', 'ПЯТНИЦА', 'СУББОТА')
        wb = openpyxl.load_workbook(filename=file_name)
        print('------------------')

        # цикл по листам в документе
        for sheet_name in [i for i in wb.sheetnames if 'ОЗФО' in i]:
            print(sheet_name)
            ws = wb[sheet_name]
            merged_ranges = ws.merged_cells.ranges
            cells_info, institute_memory = dict(), None
            self.db_obj.add_year(sheet_name)

            # цикл для получения переменных с информацией о координатах института, программы и формированию словаря с
            # днями недели
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == 'ИНСТИТУТ':
                        # Запоминаем координаты для прохождения цикла по институтам
                        inst_col, inst_row = cell.column, cell.row
                        continue
                    elif cell.value == 'ОБРАЗОВАТЕЛЬНАЯ ПРОГРАММА':
                        # и по программам
                        prog_row = cell.row
                        continue

                    elif cell.value in days_of_week:
                        cells_info[cell.value] = {'coordinate': cell.coordinate, 'time': {}}

                        for rng in merged_ranges:
                            if cell.coordinate in rng:
                                cells_info[cell.value].update({'range': rng})

                                for i in ws.iter_rows(min_row=rng.min_row, max_row=rng.max_row, min_col=rng.min_col+1,
                                                      max_col=rng.max_col+1):
                                    if i[0].value:
                                        week_odd = ws.cell(row=i[0].row, column=i[0].column + 1)
                                        week_even = ws.cell(row=i[0].row + 1, column=i[0].column + 1)

                                        cells_info[cell.value]['time'].update(
                                            {
                                                i[0].value: {
                                                    week_odd.value: week_odd.coordinate,
                                                    week_even.value: week_even.coordinate,
                                                }
                                            }
                                        )
                                break
                        continue
            pprint.pprint(cells_info)

            # цикл по институтам и программам и добавление в БД
            for col in ws.iter_cols(min_col=inst_col+1, min_row=inst_row, max_row=prog_row):
                if col[0].value:
                    institute_memory = col[0].value.replace('\n', '').capitalize()

                    if col[2].value:
                        program = col[2].value.replace('\n', '').capitalize()
                        coordinates = col[2].coordinate
                    else:
                        program = col[1].value.replace('\n', '').capitalize()
                        coordinates = col[1].coordinate

                    self.db_obj.add_inst_prog(sheet_name, institute_memory, program, coordinates)

                elif not col[0].value and (col[1].value or col[2].value):
                    if col[2].value:
                        program = col[2].value.replace('\n', '').capitalize()
                        coordinates = col[2].coordinate
                    else:
                        program = col[1].value.replace('\n', '').capitalize()
                        coordinates = col[1].coordinate

                    self.db_obj.add_prog(sheet_name, institute_memory, program, coordinates)

            # цикл по парам
            for couple in ws.iter_rows():
                pass

            print('------------------')

        # Заносим дату обновления в БД
        self.db_obj.updated()

    def __del__(self):
        self.db_obj.close_conn()


guu_obj = Guu()
